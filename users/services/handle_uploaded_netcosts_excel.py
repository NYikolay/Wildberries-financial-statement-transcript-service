import datetime
import logging
from typing import List, Set

import openpyxl

from django.db import transaction

from users.forms import ExcelNetCostsForm
from users.models import NetCost, ClientUniqueProduct


django_logger = logging.getLogger('django_logger')


def handle_uploaded_net_costs(file, current_api_key) -> dict:
    """
    The function validates the processed rows from the Excel file and saves the values to the database.
    Before creating NetCost objects, all previous ones are deleted from the database
    :param file: User uploaded Excel file
    :param current_api_key: Current Api key of current user
    :return: Returns the file processing status as a dictionary. If successful, status is True, False otherwise
    """
    excel_data: dict = get_net_costs_excel_data(file)
    if excel_data.get('status') is False:
        return excel_data

    handled_excel_data: dict = get_net_costs_form_objs_and_nm_ids(excel_data.get('excel_data'), current_api_key)

    if all([form.is_valid() for form in handled_excel_data.get('net_costs_forms')]):
        client_unique_product_objs = ClientUniqueProduct.objects.in_bulk(
            handled_excel_data.get('unique_nm_ids'),
            field_name='nm_id'
        )

        dates_validation_result: dict = validate_dates_of_uploaded_net_costs(excel_data.get('excel_data'))
        if dates_validation_result.get('status') is False:
            return dates_validation_result

        net_costs_objs_list: list = get_net_costs_objs_list(
            handled_excel_data.get('net_costs_forms'),
            client_unique_product_objs
        )
        try:
            with transaction.atomic():
                NetCost.objects.filter(
                    product__in=client_unique_product_objs.values()
                ).delete()
                NetCost.objects.bulk_create(net_costs_objs_list)
        except Exception as err:
            django_logger.critical(
                f'Failed to load net costs from excel into the database for user {current_api_key.user.email}',
                exc_info=err
            )
            return {
                'status': False,
                'message': 'Произошла ошибка во время загрузки себестоимости. Не удалось сохранить данные.'
            }
        return {
            'status': True,
            'message': 'Себестоимость успешно установлена.'
        }
    return {
        'status': False,
        'message': 'Ошиба валидации файла. Пожалуйста, убедитесь, что все значения заполнены в сответствии с шаблоном.'
    }


def handle_net_cost_obj(cleaned_data: dict, product_obj):
    return NetCost(
        product=product_obj,
        amount=cleaned_data.get('amount'),
        cost_date=cleaned_data.get('cost_date')
    )


def get_net_costs_excel_data(file) -> dict:
    """
    The function processes the Excel file taken as a parameter and, using the openpyxl library,
    generates a list containing sublists that contain the values of the user-filled lines.
    The number of columns should not exceed 3
    :param file: User uploaded Excel file
    :return: Returns the dictionary containing the status key, message if status = False or excel_data if status = True
    """
    work_book = openpyxl.load_workbook(file)
    work_book.active = 0
    worksheet = work_book.active
    excel_data: List[list] = list()

    if worksheet.max_column != 3:
        return {
            'status': False,
            'message': 'Данные не были сохранены. '
                       'Пожалуйста, убедитесь, что в файле присутствуют только следующие столбцы: '
                       'Код номенклатуры, Себестоимость, Дата начала действия.'
        }
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(cell.value)

        final_row_data_bool = map(get_validated_row_value_bool, row_data)

        if all(list(final_row_data_bool)):
            excel_data.append(row_data)

    return {
        'status': True,
        'excel_data': excel_data
    }


def get_validated_row_value_bool(value) -> bool:
    """
    The function checks if the value in the parameter is None
    :param value:
    :return: Returns False if the value is None, and True otherwise
    """
    return False if value is None else True


def get_net_costs_form_objs_and_nm_ids(excel_data: List[list], current_api_key) -> dict:
    """
    The function processes the excel_data parameter, forming a list consisting of instances
    of the ExcelNetCostsForm class (form) and a set of unique article identifiers (nm_id),
    provided that nm_id is in the current_user_nm_ids list at each iteration of the loop
    (in other words, whether nm_id belongs to the products of the current user), otherwise, the iteration is ignored
    :param excel_data: The result of the get_net_costs_excel_data function.
    :param current_api_key: Current Api key of current user
    :return: Returns a dictionary consisting of two keys.
    Instances of the ExcelNetCostsForm class and a set of unique nm_ids
    """
    current_user_nm_ids: list = ClientUniqueProduct.objects.filter(
        api_key=current_api_key
    ).values_list('nm_id', flat=True)

    net_costs_form_objs: list = list()
    unique_nm_ids: Set[int] = set()
    for nm_id, amount, cost_date in excel_data[1:]:
        if nm_id in current_user_nm_ids:
            unique_nm_ids.add(nm_id)
            net_costs_form_objs.append(ExcelNetCostsForm(
                {
                    'nm_id': nm_id,
                    'amount': amount,
                    'cost_date': cost_date
                }
            ))
    return {
        'net_costs_forms': net_costs_form_objs,
        'unique_nm_ids': list(unique_nm_ids)
    }


def get_net_costs_objs_list(handled_excel_data: list, client_unique_product_objs: dict) -> list:
    """
    The function generates a list consisting of instances of the NetCost class
    :param handled_excel_data: List of instances of the ExcelNetCostsForm class
    :param client_unique_product_objs: Dictionary containing nm_id as a key and a ClientUniqueProduct object
    :return: Returns the list of instances of the NetCost class
    """
    net_costs_objs: list = [handle_net_cost_obj(
        form.cleaned_data,
        client_unique_product_objs.get(form.cleaned_data['nm_id'])
    ) for form in handled_excel_data]
    return net_costs_objs


def validate_dates_of_uploaded_net_costs(excel_data: List[list]) -> dict:
    """
    The function checks if the dates entered by the user in the file match several conditions.
    1. The start date cannot be more than today
    2. one nm_id in the excel_data parameter cannot have two identical dates
    :param excel_data: The result of the get_net_costs_excel_data function.
    :return: Returns the dictionary with the status of the function.
    If status = False, a message is added to the dictionary to output validation error to the user
    """
    uniques = set()
    for nm_id, amount, cost_date in excel_data[1:]:
        unique_key = (nm_id, cost_date)
        if unique_key in uniques:
            return {
                'status': False,
                'message': f'Код номенклатуры - {nm_id} не может содержать две одинаковых даты.'
            }
        uniques.add(unique_key)

        if cost_date > datetime.datetime.today():
            return {
                'status': False,
                'message': f'Код номенклатуры - {nm_id}. Дата начала действия не может быть больше, чем сегодня.'
            }
    return {
        'status': True,
    }
