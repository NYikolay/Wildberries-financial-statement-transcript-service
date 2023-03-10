from openpyxl import Workbook, styles


def generate_excel_net_costs_example(net_costs_set):
    work_book = Workbook()
    columns = ['Код номенклатуры', 'Себестоимость, руб', 'Дата начала действия себестоимости, дд.мм.гггг']
    work_sheet = work_book.active
    work_sheet.title = "net_costs"
    work_sheet.merge_cells('A1:C1')
    work_sheet.column_dimensions['A'].width = 20
    work_sheet.column_dimensions['B'].width = 20
    work_sheet.column_dimensions['C'].width = 50
    work_sheet.row_dimensions[1].height = 100
    work_sheet.column_dimensions['C'].number_format = 'DD.MM.YYYY'
    work_sheet.column_dimensions['B'].number_format = '#,##0.00'
    work_sheet.column_dimensions['A'].font = styles.Font(name='Arial', bold=False, italic=False, size=10)
    work_sheet.column_dimensions['C'].font = styles.Font(name='Arial', bold=False, italic=False, size=10)
    work_sheet.column_dimensions['B'].font = styles.Font(name='Arial', bold=False, italic=False, size=10)

    cell1 = work_sheet.cell(row=1, column=1)
    cell1.value = """
    ВАЖНО! 
    Необходимо заполнять информацию строго в соответствии с шаблоном, в противном случае Себестоимость не будет установлена.
    Запрещено менять столбцы местами или менять формат ячеек.
    Значение Себестоимости не может иметь более 10 чисел до запятой и не более 2 чисел после.
    Дата начала действия Себестоимости не может быть больше, чем сегодня, также несколько значений Себестоимости для одного Кода номенклатуры 
    не могут иметь две одинаковых даты.
    Количество значений Себестоимости для одного Кода номенклатуры неограниченно.
    Соблюдать очередность Дат начала действия себестоимости не требуется.
    """
    cell1.font = styles.Font(name='Arial', color='ff0000', bold=False, italic=False, size=8)
    cell1.alignment = styles.alignment.Alignment(horizontal='left', vertical='center')

    for index, value in enumerate(columns, 1):
        cell = work_sheet.cell(row=2, column=index)
        cell.value = value
        cell.alignment = styles.alignment.Alignment(horizontal='left', vertical='center')
        cell.font = styles.Font(name='Arial', bold=True, italic=False, size=10)

    for index, row in enumerate(net_costs_set, 3):
        work_sheet[f'C{index}'].number_format = 'DD.MM.YYYY'
        work_sheet[f'B{index}'].number_format = '#,##0.00'
        for inner_index, value in enumerate(row, 1):
            cell = work_sheet.cell(row=index, column=inner_index)
            cell.value = value
            cell.alignment = styles.alignment.Alignment(horizontal='right', vertical='center')
            cell.font = styles.Font(name='Arial', bold=False, italic=False, size=10)

    return work_book
